import os
import re
import openpyxl
import pandas as pd
import warnings

warnings.simplefilter('ignore')
path = "C:\\华盛通\\技术支持周报"
os.chdir(path)
# main_workbook = openpyxl.load_workbook('线上问题处理记录汇总表20230223.xlsx')  # 保留宏，保留外部链接
# print(wb.sheetnames)
# unclosed_sheet = main_workbook['未关闭映射模板']


# print(unclosed_sheet.dimensions)


class CleanUnclosedBugWorkbook(object):
    def __init__(self, name):
        self.name = name
        self.workbook = openpyxl.load_workbook(self.name)
        self.sheet = self.workbook['Bug']
        self.modify_priority()
        self.modify_bug_status()
        self.modify_assignor()
        self.sort_date()
        self.insert_num()
        # self.delete_row_and_add_column()

    def modify_bug_status(self):
        for i in self.sheet['E']:
            if i.value == '激活':
                i.value = '处理中'
            elif i.value == '已解决':
                i.value = '验证中'
        self.workbook.save(self.name)
        print(self.name + '——修改状态完成')

    def modify_assignor(self):
        pattern = re.compile(r'\(.*?\)', flags=re.I)  # 正则表达式去除括号及括号内的内容
        for i in self.sheet['G']:
            i.value = re.sub(pattern, '', i.value)
        self.workbook.save(self.name)
        print(self.name + '——清洗指派人完成')

    def modify_priority(self):
        for i in self.sheet['H']:
            if i.value == '中':
                i.value = '普通'
        self.workbook.save(self.name)
        print(self.name + '——修改优先级别完成')

    def delete_row_and_add_column(self):
        self.sheet.delete_rows(idx=1, amount=1)  # 删除头行
        self.sheet.insert_cols(idx=1, amount=1)  # 插入头列
        for i in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row=i, column=1).value = i
        self.workbook.save(self.name)

    def sort_date(self):
        xl_file = pd.ExcelFile(self.name)
        df = xl_file.parse('Bug')
        df.sort_values(by="创建日期", ascending=True, inplace=True)
        xl_writer = pd.ExcelWriter(self.name)
        df.to_excel(xl_writer, sheet_name='Bug', index=False)
        xl_writer.close()
        print(self.name + '——排序改成升序完成')

    # 插入序号到第一列
    def insert_num(self):
        if not self.sheet['A1'].value == 'No':
            self.sheet.insert_cols(idx=1, amount=1)  # 在第一列左侧插入1列
            self.sheet.cell(row=1, column=1, value='No')
            for i in range(2, self.sheet.max_row + 1):
                self.sheet.cell(row=i, column=1, value=i - 1)
            self.workbook.save(self.name)
            print(self.name + '——插入序号完成')
        else:
            print(self.name + '——不需要插入序号！！！')


class CleanAllBugWorkbook(CleanUnclosedBugWorkbook):
    def __init__(self, name):
        super().__init__(name)
        self.j = 0
        self.modify_priority()
        self.modify_bug_status()
        self.modify_assignor()
        self.fill_assignor()
        self.sort_date()

    def modify_bug_status(self):
        for i in self.sheet['F']:  # F不一样
            if i.value == '激活':
                i.value = '处理中'
            elif i.value == '已解决':
                i.value = '验证中'
        self.workbook.save(self.name)
        print(self.name + '——修改状态完成')

    def modify_priority(self):
        for i in self.sheet['C']:
            if i.value == '中':
                i.value = '普通'
        self.workbook.save(self.name)
        print(self.name + '——修改优先级别完成')

    def fill_assignor(self):
        for i in range(len(self.sheet['G'])):
            if self.sheet.cell(row=i + 1, column=7).value == 'Closed':
                self.sheet.cell(row=i + 1, column=7).value = self.sheet.cell(row=i + 1, column=10).value
        self.workbook.save(self.name)


class CleanWorkOrderBook(object):
    def __init__(self, name1, name2, name3):
        self.name1 = name1
        self.name2 = name2
        self.name3 = name3
        self.workbook1 = openpyxl.load_workbook(self.name1)
        self.workbook2 = openpyxl.load_workbook(self.name2)
        self.workbook3 = openpyxl.load_workbook(self.name3)
        self.sheet1 = self.workbook1['第一页']
        self.sheet2 = self.workbook2['第一页']
        self.sheet3 = self.workbook3['第一页']
        self.add_up()
        self.clean_consult_content()

    def add_up(self):
        data0 = []  # 复制表头数据
        wb0 = openpyxl.load_workbook(filename=self.name1)
        ws0 = wb0.active
        for i in range(2, ws0.max_column):  # 从1开始，到最后一列（因为工单表格有缺陷，所以不用+1，下同）
            data0.append(ws0.cell(row=2, column=i).value)  # data0追加数据，如工单编号、华盛号、顾客姓名等，废弃掉第一列数据
        data1 = []
        name_list = [self.name1, self.name2, self.name3]
        for n in range(len(name_list)):
            name = name_list[n]
            wb1 = openpyxl.load_workbook(filename=name)  # 依次读取三个表
            ws1 = wb1.active
            for i in range(3, ws1.max_row + 1):  # 从第3行起到最后一行。第一行写完再开始下一行的循环。
                list = []
                for j in range(2, ws1.max_column):  # 从第1列到最后一列
                    list.append(ws1.cell(row=i, column=j).value)  # 把每个单元格数据添加到第一个列表，['','','']
                data1.append(list)  # 把每个表的数据添加到第二个列表[[],[],[]]

        data = []
        data.append(data0)  # 添加表头
        for k in range(len(data1)):  # 添加数据
            data.append(data1[k])
        wb = openpyxl.Workbook()  # 新建工作簿
        ws = wb.active  # 新建工作表
        ws.title = '汇总'
        for n_row in range(1, len(data) + 1):  # 写入数据，从第一行到最后一行
            for n_col in range(1, len(data[n_row - 1]) + 1):
                ws.cell(row=n_row, column=n_col, value=str(data[n_row - 1][n_col - 1]))
        wb.save(filename='总表.xlsx')  # 保存xlsx
        print('汇总完成')

    def clean_consult_content(self):
        list = [2, 3, 10]
        LIST = [[], [], []]
        new_LIST = ['咨询内容']
        workbook = openpyxl.load_workbook('总表.xlsx')
        sheet = workbook.active
        sheet.insert_cols(idx=5, amount=1)
        for k in range(len(list)):
            for i in range(1, sheet.max_row + 1):
                LIST[k].append(sheet.cell(row=i, column=list[k]).value)
        for j in range(0, len(LIST[0]) - 1):  # LIST[0]长度只有69，下一行加1后，在循环最大时，索引会变成70，就超出索引了。
            new_LIST.append(str(LIST[0][j + 1]) + ' ' + str(LIST[1][j + 1]) + ' ' + str(LIST[2][j + 1]))
        for n in range(0, sheet.max_row):  # 写入数据
            sheet.cell(row=n + 1, column=5, value=new_LIST[n])
        col_num = [5, 6, 7, 8, 11, 13, 12]
        new_sheet = workbook.create_sheet('新')
        for r in range(0, len(col_num)):
            for m in range(1, sheet.max_row + 1):
                new_sheet.cell(row=m, column=r + 1).value = sheet.cell(row=m,
                                                                       column=col_num[r]).value  # [列，行] = [列的切片值，行]
        workbook.save('总表.xlsx')


clean_unclosed_bug = CleanUnclosedBugWorkbook('线上问题记录-未关闭Bug.xlsx')
clean_all_bug = CleanAllBugWorkbook('线上问题记录-所有Bug.xlsx')
cleanworkorderbook = CleanWorkOrderBook('工单查询 (1).xlsx', '工单查询 (2).xlsx', '工单查询 (3).xlsx')

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.utils import get_column_letter


def xlsx_sheet_copy(src_path, tag_path, sheet_name):  # 跨xlsx复制sheet
    # 跨xlsx文件复制源文件xlsx中指定的sheet
    # 保留所有格式，以及行高列宽，视觉效果几乎一致
    # 不能复制除了文字以外的东西，例如图片
    # src_path:源xlsx文件路径
    # tag_path:目标xlsx文件路径
    # sheet_name:需要复制的源xlsx文件sheet的名称
    src_workbook = load_workbook(src_path)  # 打开源xlsx
    src_file_sheet = src_workbook[sheet_name]  # 打开目标sheet
    tag_workbook = load_workbook(tag_path)  # 打开目标xlsx
    tag_file_sheet = tag_workbook.create_sheet(sheet_name)  # 新建一个同名空sheet等待写入

    for row in src_file_sheet:
        # 遍历源xlsx文件制定sheet中的所有单元格
        for cell in row:  # 复制数据
            tag_file_sheet[cell.coordinate].value = cell.value
            if cell.has_style:  # 复制样式
                tag_file_sheet[cell.coordinate].font = copy(cell.font)
                tag_file_sheet[cell.coordinate].border = copy(cell.border)
                tag_file_sheet[cell.coordinate].fill = copy(cell.fill)
                tag_file_sheet[cell.coordinate].number_format = copy(
                    cell.number_format
                )
                tag_file_sheet[cell.coordinate].protection = copy(cell.protection)
                tag_file_sheet[cell.coordinate].alignment = copy(cell.alignment)

    wm = list(zip(src_file_sheet.merged_cells))  # 开始处理合并单元格
    if len(wm) > 0:  # 检测源xlsx中合并的单元格
        for i in range(0, len(wm)):
            cell2 = (
                str(wm[i]).replace("(<MergedCellRange ", "").replace(">,)", "")
            )  # 获取合并单元格的范围
            tag_file_sheet.merge_cells(cell2)  # 合并单元格
    # 开始处理行高列宽
    for i in range(1, src_file_sheet.max_row + 1):
        tag_file_sheet.row_dimensions[i].height = src_file_sheet.row_dimensions[
            i
        ].height
    for i in range(1, src_file_sheet.max_column + 1):
        tag_file_sheet.column_dimensions[
            get_column_letter(i)
        ].width = src_file_sheet.column_dimensions[get_column_letter(i)].width

    tag_workbook.save(tag_path)  # 保存
    tag_workbook.close()  # 关闭文件
    src_workbook.close()


xlsx_sheet_copy('C:\华盛通\技术支持周报\线上问题记录-未关闭Bug.xlsx', 'C:\华盛通\技术支持周报\线上问题处理记录汇总表20230223.xlsx', 'Bug')

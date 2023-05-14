import os
import openpyxl
import xlsxwriter

list = []
path = 'C:\华盛通\技术支持周报'
os.chdir(path)
wb = openpyxl.load_workbook('线上问题处理记录汇总表20230223.xlsx')
# print(wb.sheetnames)
unclosed_sheet = wb['未关闭映射模板']
# print(unclosed_sheet.dimensions)
cell = unclosed_sheet['A1:I28']
for i in cell:
    for j in i:
        list.append(str(j.value))
    list.append('\n')
new_wb = xlsxwriter.Workbook('new.xlsx')  # 创建新的工作簿
worksheet = new_wb.add_worksheet('未关闭映射模板')
for n in range(28):         # 换行
    for i in range(len(list)):
        if list[i] == '\n':
            break
        worksheet.write(n, i, list[i+n*10])
new_wb.close()
